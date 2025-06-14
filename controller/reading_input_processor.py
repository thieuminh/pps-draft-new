import platform
import time
from datetime import datetime
from model.Logger import Logger
import config
from config import bcolors
from model.hallway_simulator_module.HallwaySimulator import DirectoryManager
from openpyxl import load_workbook
import math
import numpy as np
import pdb
from controller.start_node_generator import StartNodeGenerator
    
#Sẽ được lớp TsgFileEditor kế thừa
class ReadingInputProcessor(StartNodeGenerator):
    def __init__(self, _dm):
        super().__init__()
        self.logger = Logger()
        self._print_out = True
        #self.x = {}
        #self.y = {}

        #config.count = 0  # reset bộ đếm
        self.dm = _dm
        if self.dm is not None:
            self.dm.full_cleanup()  # dọn sạch ban đầu
        self.processed_numbers = []
        self._space_edges = []
        self._draw = 0
        self._d = 0
        self._num_max_agvs = 0

    # Getter và Setter cho print_out
    @property
    def print_out(self):
        return self._print_out

    @print_out.setter
    def print_out(self, value):
        if not isinstance(value, bool):
            raise ValueError("print_out must be a boolean")
        self._print_out = value
    
    # Getter and Setter for d
    @property
    def d(self):
        return self._d

    @d.setter
    def d(self, value):
        self._d = value
        
    @property
    def draw(self):
        return self._draw
    @draw.setter
    def draw(self, value):
        self._draw = value

    # Getter và Setter cho space_edges
    @property
    def space_edges(self):
        return self._space_edges

    @space_edges.setter
    def space_edges(self, value):
        if not isinstance(value, list):
            raise ValueError("space_edges must be a list")
        self._space_edges = value
        
    # Getter và Setter cho num_max_agvs
    @property
    def num_max_agvs(self):
        return self._num_max_agvs

    @num_max_agvs.setter
    def num_max_agvs(self, value):
        self._num_max_agvs = value
    
    def _process_number(self, num):
        import math
        if num < 5:
            return 0
        else:
            return math.ceil(num)
    
    def get_os(self):
        os_name = platform.system()
        if os_name == 'Darwin':
            return "macOS"
        elif os_name == 'Windows':
            return "Windows"
        elif os_name == 'Linux':
            return "Linux"
        else:
            return "Undefined OS"

    def choose_solver(self):
        print("Choose the method for solving:")
        print("1 - Use LINK II solver")
        print("2 - Use parallel network-simplex")
        print("3 - Use NetworkX")
        choice = 3
        if(config.count % 2 == 0):
            choice = 1
            config.solver_choice = 'solver'
        else:
            if(config.count <= 1):
                choice = input("Enter your choice (1 or 2 or 3): ")
                if choice == '1':
                    config.solver_choice = 'solver'
                elif choice == '2':
                    config.solver_choice = 'network-simplex'
                elif choice == '3':
                    config.solver_choice = 'networkx'
                else:
                    print("Invalid choice. Defaulting to Network X.")
                    config.solver_choice = 'networkx'
            else:
                config.solver_choice = 'networkx'

    def choose_time_measurement(self):
        if(config.count == 1 and config.test_automation == 0):
            print("Choose level of Time Measurement:")
            print("0 - Fully Random")
            print("1 - Random in a list")
            print("2 - SFM")
            choice = input("Enter your choice (0 to 2): ")
            if choice == '0':
                config.level_of_simulation = 0
            elif choice == '1':
                config.level_of_simulation = 1
            elif choice == '2':
                config.level_of_simulation = 2
            else:
                print("Invalid choice. Defaulting to run SFM.")
                config.level_of_simulation = 2
        else:
            if(config.count <= 2):
                config.level_of_simulation = 0
            elif(config.count <= 4):
                config.level_of_simulation = 1
            elif(config.count <= 6):
                config.level_of_simulation = 2
        if(config.level_of_simulation == 1):
            #random in the list
            self.read_xls()

    def choose_test_automation(self):
        if(config.count == 1):
            print("Choose level of Test automation:")
            print("0 - Manual")
            print("1 - Automation")
            choice = input("Enter your choice (0 or 1): ")
            if choice == '0':
                config.test_automation = 0
            else:
                print("Defaulting to run Automation")
                config.test_automation = 1

    def start_round(self):
        config.count += 1
    
        now = datetime.now()
        dt_string = now.strftime("%d/%m/%Y %H:%M:%S")
        print(f"----- ROUND {config.count} at {dt_string} -----")
    
        # Cảnh báo nếu không hỗ trợ mô phỏng SFM
        if config.count >= 5 and self.get_os() != 'Linux':
            print("⚠️  The current OS doesn't support SFM Simulation")
            return False  # báo hiệu không nên tiếp tục vòng lặp
    
        # In cảnh báo half cleanup từ vòng lặp 2 trở đi
        if config.count > 1:
            print(f"{bcolors.WARNING}Start half cleanup{bcolors.ENDC}")
    
        # Thông báo chọn solver
        if config.count % 2 == 0:
            print("Start using solver at:", config.count)
        else:
            print("Start using NetworkX at:", config.count)
    
        # Cleanup tạm
        if self.dm is not None:
            self.dm.half_cleanup()
        #time.sleep(1)
    
        return True  # báo hiệu tiếp tục vòng lặp
    
    def read_xls(self):
        # Đọc file Excel
        file_name = 'completion_times.xlsx'
        workbook = load_workbook(file_name, data_only=True)
        sheet = workbook.active
        
        # Find the last column with a value in the first row
        last_column_with_value = sheet.max_column
        for col in range(sheet.max_column, 0, -1):
            if sheet.cell(row=1, column=col).value is not None:
                last_column_with_value = col
                break
        # Lấy số lượng cột và hàng
        max_column, max_row = sheet.max_column, sheet.max_row
        # Lấy dữ liệu từ 3 cột cuối cùng
        last_three_columns = []
        for row in sheet.iter_rows(min_row=1, max_row=max_row, \
            min_col=last_column_with_value-2, max_col=last_column_with_value):
            row_data = []
            for cell in row:
                if cell.value is not None:
                    row_data.append(cell.value)
            if row_data:
                last_three_columns.append(row_data)
        
        for row in last_three_columns:
            for num in row:
                if isinstance(num, (int, float)):  # Kiểm tra nếu là số
                    self.processed_numbers.append(self._process_number(num))
    
    def ask_for_print_out(self, use_config_data = False):
        if(use_config_data):
            self.print_out = config.print_output
        else:
            print_out = input("Bạn có muốn print out ra hết các thông báo chi tiết khi chương trình hoạt động không? (Enter để trả lời KHÔNG): ")
            if print_out == '':
                self.print_out = False
            else:
                self.print_out = True
        config.print_output = self.print_out
        
    def ask_spatial_map(self, use_config_data = False):
        if(use_config_data):
            filepath = config.filepath
        else:
            filepath = input("Nhap ten file can thuc hien (hint: Redundant3x3Wards.txt): ")
            if filepath == '':
                filepath = 'Redundant3x3Wards.txt'
            config.filepath = filepath
            
    def handle_n(self, parts):
        nid, ntype = int(parts[1]), parts[2]
        if ntype == '1':
            self.started_nodes.append(nid)
        elif ntype == '-1':
            self.ID.append(nid)
            self.earliness = [] if isinstance(self.earliness, int) else self.earliness
            self.tardiness = [] if isinstance(self.tardiness, int) else self.tardiness
            self.earliness.append(int(parts[3]))
            self.tardiness.append(int(parts[4]))
            
    def process_input_file(self, filepath):
        self.space_edges = []
        try:
            with open(filepath, 'r') as file:
                self.M = 0
                for line in file:
                    parts = line.strip().split()
                    if not parts:
                        continue
                    tag = parts[0]
                    if tag == 'a' and len(parts) >= 6:
                        id1, id2 = int(parts[1]), int(parts[2])
                        self.space_edges.append(parts)
                        self.M = max(self.M, id1, id2)
                    elif tag == 'n':
                        self.handle_n(parts)
                    elif tag == 'alpha':
                        self.alpha = int(parts[1])
                    elif tag == 'beta':
                        self.beta = int(parts[1])
            config.M = self.M
            if self.print_out:
                print("Doc file hoan tat, M =", self.M)
        except FileNotFoundError:
            if self.print_out:
                print("File khong ton tai!")
                
    def ask_horizontal_time(self, use_config_data = False):
        if(use_config_data):
            self.H = config.H
        else:
            self.H = input("Nhap thoi gian can gia lap (default: 10): ")
            if(self.H == ''):
                self.H = 10
            else:
                self.H = int(self.H)
            config.H = self.H
            
    def ask_for_draw(self, use_config_data = False):
        if(use_config_data):
            self.draw = config.draw
        else:
            self.draw = input("Nhập 1 để vẽ TSG (default 0) không nên dùng với đồ thị lớn): ")
            if(self.draw == '' or self.draw == 0):
                self.draw = 0
            else:
                self.draw = 1
            config.draw = self.draw
            
    def ask_for_d(self, use_config_data = False):
        if(use_config_data):
            self.d = config.d
        else:
            self.d = input("Nhap time unit (default: 10): ")
            if(self.d == ''):
                self.d = 10
            else:
                self.d = int(self.d)
            config.d = self.d
            
    def generate_time_windows(self):
        self.num_max_agvs = input("Nhap so luong AGV toi da di chuyen trong toan moi truong (default: 2):")
        if(self.num_max_agvs == ''):
            self.num_max_agvs = 2
        else:
            self.num_max_agvs = int(self.num_max_agvs)
        num_of_agvs = self.num_max_agvs
        config.num_max_agvs = self.num_max_agvs
        config.numOfAGVs = num_of_agvs
        self.generate_start_nodes(num_of_agvs)
        return num_of_agvs
    
    def generate_new_agvs(self):
        num_of_additional_agvs = config.numOfAGVs - len(config.ID)
        for _ in range(num_of_additional_agvs):
            [s, d, e, t] = self.generate_numbers_student(self.M, self.H, 12, 100)
            while(d in self.ID or s in self.started_nodes):
                [s, d, e, t] = self.generate_numbers_student(self.M, self.H, 12, 100)
            self.started_nodes.append(s)
            self.ID.append(d)
            config.ID.append(d)
            self.earliness.append(e)
            self.tardiness.append(t)
    def reuse_for_tasks(self, use_config_data = False):
        num_of_agvs = 0
        if(use_config_data):
            self.num_max_agvs = config.num_max_agvs
            self.ID = config.ID
            self.earliness = config.earliness
            self.tardiness = config.tardiness
            for i in range(len(config.started_nodes)):
                if (config.started_nodes[i] > self.M and config.started_nodes[i] % self.M not in config.started_nodes):
                    config.started_nodes[i] = config.started_nodes[i] % self.M
            self.started_nodes = config.started_nodes
            num_of_agvs = config.numOfAGVs
            if(config.numOfAGVs > len(config.ID)):
                self.generate_new_agvs()
            elif(config.numOfAGVs < len(config.ID)):
                config.ID = config.ID[:(config.numOfAGVs)]
                config.earliness = config.earliness[:(config.numOfAGVs)]
                config.tardiness = config.tardiness[:(config.numOfAGVs)]
                config.started_nodes = config.started_nodes[:(config.numOfAGVs)]
                self.earliness = self.earliness[:(config.numOfAGVs)]
                self.tardiness = self.tardiness[:(config.numOfAGVs)]
                self.started_nodes = self.started_nodes[:(config.numOfAGVs)]
        else:
            num_of_agvs = self.generate_time_windows()
        return num_of_agvs


